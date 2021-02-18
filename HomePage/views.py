import openpyxl
from django.contrib import auth
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from openpyxl import load_workbook

from .forms import RegistrationForm
from .forms import LoginForm
from .forms import UploadFileForm
from .models import handle_uploaded_file
from .models import Customer
from .forms import DocumentForm
from django.shortcuts import redirect
from pyexcel_xlsx import get_data as xlsx_get
from django.utils.datastructures import MultiValueDictKeyError


# Create your views here.
def index(request):
    # response = HttpResponse()
    # response.writelines('<h1>Xin chào</h1>')
    # response.write('Đây là app home')
    # return response
    return render(request, 'pages/home.html')


# Create your views here.
def contact(request):
    return render(request, 'pages/contact.html')


def error404(request, exception):
    return render(request, exception, 'pages/error.html')


def error500(request):
    return render(request, 'pages/error.html')


def register(request):
    form = RegistrationForm()
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/home')
    return render(request, 'pages/register.html', {'form': form})


def user_login(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = auth.authenticate(username=cd['username'],
                                     password=cd['password'])
            if user is not None:
                if user.is_active:
                    auth.login(request, user)
                    return HttpResponse('Authenticated successfully')
                else:
                    return HttpResponse('Disable account')
            else:
                return HttpResponse('Invalid account')
    else:
        form = LoginForm()
    return render(request, 'pages/login.html', {'form': form})


def luckynum(request):
    return render(request, 'pages/QuaySo.html')


def logout():
    return None


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.FILES['file'])
            return HttpResponseRedirect('/success/url')
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})


from .models import import_data


def simple_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        test = import_data(uploaded_file_url)
        print(test)
        return render(request, 'page/simple_upload.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'pages/simple_upload.html')


def model_form_upload(request):
    if "GET" == request.method:
        return render(request, 'pages/model_form_upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        return render(request, 'pages/model_form_upload.html', {"excel_data": excel_data})

def reponse_winner(request):
    pass